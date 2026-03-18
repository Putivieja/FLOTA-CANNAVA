#!/usr/bin/env python3
"""
generate.py — Cannava Fleet Dashboard Generator
Lee el Excel de flota y genera index.html con todos los datos embebidos.
"""

import openpyxl
import json
import re
import glob
import sys
from datetime import datetime

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def norm(s):
    """Normaliza texto para comparar nombres de columnas del Excel."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    # Quitar tildes y ñ
    for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N"),
                 ("À","A"),("È","E"),("Ì","I"),("Ò","O"),("Ù","U")]:
        s = s.replace(a, b)
    # Reemplazar cualquier caracter no alfanumérico por _
    s = re.sub(r'[^A-Z0-9]', '_', s)
    # Colapsar múltiples _ y sacar los de los extremos
    s = re.sub(r'_+', '_', s).strip('_')
    return s

def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()

def safe_num(v):
    try:
        return float(v) if v not in (None, "", "-") else None
    except:
        return None

def safe_int(v, default=0):
    try:
        return int(float(v)) if v not in (None, "", "-") else default
    except:
        return default

def to_bool_estado(v):
    """HECHO → True | EN PROCESO → None | PENDIENTE → False"""
    s = safe_str(v).upper()
    if "HECHO" in s or s in ("TRUE", "1", "SI", "SÍ"):
        return True
    if "PROCESO" in s or "PROCESS" in s:
        return None
    return False

def get_sheet(wb, *candidates):
    """Busca una hoja por nombre (tolerante a mayúsculas/minúsculas)."""
    names_map = {s.upper().strip(): s for s in wb.sheetnames}
    for c in candidates:
        key = c.upper().strip()
        if key in names_map:
            return wb[names_map[key]]
    # Fallback: buscar por substring
    for c in candidates:
        for name in wb.sheetnames:
            if c.upper() in name.upper():
                return wb[name]
    return None

def sheet_to_dicts(sheet):
    """Convierte hoja de Excel a lista de dicts.
    Busca la fila real de headers ignorando filas de titulo."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    # Buscar la fila que contiene PATENTE o DOMINIO como header real
    header_idx = 0
    for i, row in enumerate(rows):
        row_vals = [norm(str(v)) for v in row if v is not None]
        if any(k in row_vals for k in ["PATENTE", "DOMINIO_PATENTE", "DOMINIO__PATENTE_", "FECHA"]):
            if any(k in row_vals for k in ["PATENTE", "DOMINIO_PATENTE", "DOMINIO__PATENTE_"]):
                header_idx = i
                break
            elif "FECHA" in row_vals and i > 0:
                header_idx = i
                break
    headers = [norm(h) for h in rows[header_idx]]
    result = []
    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        result.append(d)
    return result

def get(d, *keys):
    """Busca el primer valor no vacío entre las claves dadas."""
    for k in keys:
        nk = norm(k)
        if nk in d and d[nk] not in (None, "", "-"):
            return d[nk]
    return None

# ═══════════════════════════════════════════════════════════════
# LEER EL EXCEL
# ═══════════════════════════════════════════════════════════════

# Buscar el archivo .xlsx en la carpeta raíz
xlsx_files = glob.glob("*.xlsx") + glob.glob("*.XLSX")
if not xlsx_files:
    print("❌ ERROR: No se encontró ningún archivo .xlsx en la carpeta")
    sys.exit(1)

# Si hay varios, toma el primero que tenga "FLOTA" en el nombre, sino el primero
flota_files = [f for f in xlsx_files if "FLOTA" in f.upper()]
xlsx_path = flota_files[0] if flota_files else xlsx_files[0]
print(f"📂 Leyendo: {xlsx_path}")

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print(f"📋 Hojas encontradas: {wb.sheetnames}")

# Obtener las hojas
sh_flota  = get_sheet(wb, "FLOTA")
sh_estado = get_sheet(wb, "ESTADO DE FLOTA", "ESTADO_DE_FLOTA", "ESTADO")
sh_comb   = get_sheet(wb, "COMBUSTIBLE")

if not sh_flota:
    print("❌ ERROR: No se encontró la hoja 'FLOTA'")
    sys.exit(1)

flota_rows  = sheet_to_dicts(sh_flota)
estado_rows = sheet_to_dicts(sh_estado) if sh_estado else []
comb_rows   = sheet_to_dicts(sh_comb)   if sh_comb  else []

print(f"✓ FLOTA: {len(flota_rows)} vehículos")
print(f"✓ ESTADO DE FLOTA: {len(estado_rows)} reparaciones")
print(f"✓ COMBUSTIBLE: {len(comb_rows)} registros")

# DEBUG: mostrar columnas detectadas
if flota_rows:
    print(f"  Columnas FLOTA: {list(flota_rows[0].keys())[:8]}")
if estado_rows:
    print(f"  Columnas ESTADO: {list(estado_rows[0].keys())[:8]}")
if comb_rows:
    print(f"  Columnas COMBUSTIBLE: {list(comb_rows[0].keys())[:6]}")

# ═══════════════════════════════════════════════════════════════
# AGREGAR COMBUSTIBLE POR PATENTE
# ═══════════════════════════════════════════════════════════════

comb_by_pat = {}
for r in comb_rows:
    # La columna puede llamarse DOMINIO/PATENTE o PATENTE
    pat_raw = get(r, "DOMINIO_PATENTE_", "DOMINIO_PATENTE", "DOMINIO__PATENTE_", "PATENTE", "DOMINIO")
    litros_raw = get(r, "LITROS")
    if not pat_raw or not litros_raw:
        continue
    pat = safe_str(pat_raw).upper().replace(" ", "").strip()
    litros = safe_num(litros_raw)
    if pat and litros and litros > 0:
        if pat not in comb_by_pat:
            comb_by_pat[pat] = {"litros": 0.0, "cargas": 0}
        comb_by_pat[pat]["litros"] += litros
        comb_by_pat[pat]["cargas"] += 1

# ═══════════════════════════════════════════════════════════════
# AGREGAR REPARACIONES POR PATENTE
# ═══════════════════════════════════════════════════════════════

reps_by_pat = {}
for r in estado_rows:
    pat = safe_str(get(r, "PATENTE") or "").upper().replace(" ", "").strip()
    if not pat:
        continue

    tipo = safe_str(get(r, "TIPO_DE_REPARACION", "TIPO_DE_REPARACION_", "TIPO_REPARACION", "TIPO_DE_REPARACI_N", "TIPO") or "")
    if not tipo:
        continue

    estado_val = get(r, "ESTADO") or "PENDIENTE"

    # Costos — intentar distintas variantes del nombre de columna
    costo_rep = safe_num(get(r, "PRESUPUESTO___", "PRESUPUESTO", "REPUESTO_1", "REPUESTO1"))
    costo_mo  = safe_num(get(r, "M_O_1___", "M_O_1__", "M_O_1", "MO_1"))
    prov1     = safe_str(get(r, "PROVEEDOR") or "")
    prov2     = safe_str(get(r, "PROVEEDOR_1") or "")

    rep = {
        "tipo":          tipo,
        "realizado":     to_bool_estado(estado_val),
        "costoRepuesto": costo_rep,
        "costoMO":       costo_mo,
        "proveedor":     prov1,
        "proveedor2":    prov2,
        "fechaProgramada": "",
        "observaciones": safe_str(get(r, "OBSERVACIONES") or ""),
    }

    if pat not in reps_by_pat:
        reps_by_pat[pat] = []
    reps_by_pat[pat].append(rep)

# ═══════════════════════════════════════════════════════════════
# CONSTRUIR EL ARRAY FLEET
# ═══════════════════════════════════════════════════════════════

FLEET = []
for i, v in enumerate(flota_rows, 1):
    pat = safe_str(get(v, "PATENTE") or "").upper().replace(" ", "").strip()
    if not pat:
        continue

    # Combustible acumulado
    comb_data = None
    if pat in comb_by_pat:
        c = comb_by_pat[pat]
        comb_data = {
            "litros": round(c["litros"], 1),
            "cargas": c["cargas"]
        }

    # Estado %
    estado_pct = safe_int(get(v, "ESTADO %", "ESTADO_%", "ESTADO_PCT", "ESTADO"), 50)
    estado_pct = max(0, min(100, estado_pct))

    # Requiere batería — acepta SI/NO, X, TRUE/FALSE
    bat_raw = safe_str(get(v, "REQUIERE BATERIA", "REQUIERE_BATERIA", "REQUIERE BATERÍA") or "").upper()
    req_bat = bat_raw in ("SI", "SÍ", "YES", "TRUE", "1", "X", "✓", "S")

    # Cubiertas
    cub_cambiar     = safe_int(get(v, "CUBIERTAS A CAMBIAR", "CUBIERTAS_A_CAMBIAR", "CUBIERTAS"), 0)
    cub_reemplazadas = safe_int(get(v, "CUBIERTAS REEMPLAZADAS", "CUBIERTAS_REEMPLAZADAS"), 0)

    # Año — la ñ puede perderse en distintas codificaciones
    anio = safe_int(get(v, "AÑO", "ANO", "AO", "YEAR"), 2020)

    entry = {
        "id":                   i,
        "vehiculo":             safe_str(get(v, "MODELO", "VEHICULO", "VEHÍCULO", "DESCRIPCION") or ""),
        "patente":              pat,
        "marca":                safe_str(get(v, "MARCA") or ""),
        "tipo":                 safe_str(get(v, "TIPO") or "CAMIONETA").upper(),
        "año":                  anio,
        "chasis":               safe_str(get(v, "Nº CHASIS", "N_CHASIS", "CHASIS", "NRO_CHASIS") or ""),
        "motor":                safe_str(get(v, "Nº MOTOR", "N_MOTOR", "MOTOR", "NRO_MOTOR") or ""),
        "responsable":          safe_str(get(v, "RESPONSABLE") or ""),
        "estadoPct":            estado_pct,
        "requiereBateria":      req_bat,
        "rodado":               safe_str(get(v, "RODADO") or ""),
        "cubiertasACambiar":    cub_cambiar,
        "cubiertasReemplazadas": cub_reemplazadas,
        "fechaCambioCubiertas": None,
        "estetica":             safe_str(get(v, "ESTETICA", "ESTÉTICA", "ESTETICA_DEP_JUDICIAL") or ""),
        "observaciones":        safe_str(get(v, "OBSERVACIONES") or ""),
        "combustible":          comb_data,
        "reparaciones":         reps_by_pat.get(pat, []),
    }
    FLEET.append(entry)

print(f"✓ Vehículos procesados: {len(FLEET)}")

# ═══════════════════════════════════════════════════════════════
# GENERAR index.html
# ═══════════════════════════════════════════════════════════════

# Leer template
try:
    with open("template.html", "r", encoding="utf-8") as f:
        html = f.read()
except FileNotFoundError:
    print("❌ ERROR: No se encontró template.html")
    sys.exit(1)

# Serializar FLEET a JSON
fleet_json = json.dumps(FLEET, ensure_ascii=False, indent=2)

# Inyectar en el placeholder
html = html.replace(
    "const FLEET = /*__FLEET_DATA__*/[];",
    f"const FLEET = {fleet_json};"
)

# Escribir index.html
with open("index.html", "w", encoding="utf-8") as f:
    f.write(html)

fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
print(f"✅ index.html generado correctamente — {fecha}")
print(f"   Vehículos: {len(FLEET)}")
print(f"   Reparaciones: {sum(len(v['reparaciones']) for v in FLEET)}")
comb_total = sum(v['combustible']['litros'] for v in FLEET if v['combustible'])
print(f"   Combustible total: {comb_total:,.0f} L")
